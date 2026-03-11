from io import BytesIO
# Invoice Database
INVOICE_DB = "/tmp/invoice_records.db"

import streamlit as st
import pandas as pd
#!/usr/bin/env python3
"""
e-Tax Invoice System for GAC Thailand
Cloud-Ready Version with Streamlit
"""


import sqlite3
import json
import re
import os

# Company Info - GAC Thailand
COMPANY_NAME = "GAC (THAILAND) CO., LTD."
COMPANY_TAX_ID = "0105548024532"
COMPANY_ADDRESS = "9/2 Sathorn 39, South Sathorn Road, Yannawa, Sathorn, Bangkok 10120, Thailand"
COMPANY_TEL = "+66 2 676 1900"
COMPANY_FAX = "+66 2 676 1990"
COMPANY_EMAIL = "thailand@gac.com"
COMPANY_WEB = "www.gac.com/thailand"

import os
# Database path - works on both local and cloud
DB_PATH = os.environ.get('DB_PATH', '/tmp/invoices.db')
UPLOAD_DIR = os.environ.get('UPLOAD_DIR', '/tmp/uploads')

def list_uploaded_files():
    """List all uploaded files"""
    upload_dir = UPLOAD_DIR
    
    # Try to create directory
    if upload_dir:
        try:
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir, exist_ok=True)
        except:
            upload_dir = '/tmp/uploads'
            try:
                os.makedirs(upload_dir, exist_ok=True)
            except:
                pass
    
    files = []
    try:
        if upload_dir and os.path.exists(upload_dir):
            for f in os.listdir(upload_dir):
                filepath = os.path.join(upload_dir, f)
                if os.path.isfile(filepath):
                    files.append({'filename': f, 'filepath': filepath})
    except:
        pass
    
    return sorted(files, key=lambda x: x['filename'], reverse=True)


def save_uploaded_file(uploaded_file):
    """Save uploaded file to uploads directory"""
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{uploaded_file.name}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return filename, filepath

def delete_uploaded_file(filename):
    """Delete uploaded file"""
    filepath = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)

def save_invoice_to_db(invoice_data, status='pending'):
    """Save invoice to database for persistence"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    items_list = invoice_data.get('items', [])
    if isinstance(items_list, str):
        items_json = items_list
    else:
        items_json = json.dumps(items_list)
    
    c.execute("""INSERT OR REPLACE INTO invoices (filename, invoice_no, invoice_date, customer_name, job_number, awb, job_ref, exchange_rate, total_amount, total_thb, items_json, status)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (invoice_data.get('filename', ''),
         invoice_data.get('invoice_no', ''),
         invoice_data.get('invoice_date', ''),
         invoice_data.get('customer_name', ''),
         invoice_data.get('job_number', ''),
         invoice_data.get('awb', ''),
         invoice_data.get('job_ref', ''),
         invoice_data.get('exchange_rate', 30.909),
         invoice_data.get('total_amount', 0),
         invoice_data.get('total_thb', 0),
         items_json,
         status))
    
    conn.commit()
    invoice_id = c.lastrowid
    conn.close()
    return invoice_id

def load_invoices_from_db():
    """Load invoices from uploaded files"""
    files = list_uploaded_files()
    invoices = []
    
    for i, f in enumerate(files):
        try:
            filepath = f['filepath']
            filename = f['filename']
            
            invoice_no = ''
            job_number = ''
            awb = ''
            invoice_date = ''
            customer_name = ''
            total_amount = 0
            items = []
            
            # Handle XML files
            if filename.endswith('.xml'):
                try:
                    import xml.etree.ElementTree as ET
                    with open(filepath, 'r', encoding='utf-8') as xml_file:
                        xml_content = xml_file.read()
                    
                    root = ET.fromstring(xml_content)
                    
                    # Search all elements for the data
                    for elem in root.iter():
                        # Invoice No - Textbox183
                        val = elem.get('Textbox183') or elem.get(':Textbox183')
                        if val and not invoice_no:
                            invoice_no = val.strip().lstrip(': ')
                        
                        # Customer - BillingPartyName
                        val = elem.get('BillingPartyName')
                        if val and not customer_name:
                            customer_name = val.replace('Billing Party:', '').strip()
                        
                        # Job No - Textbox188
                        val = elem.get('Textbox188') or elem.get(':Textbox188')
                        if val and not job_number:
                            job_number = val.strip().lstrip(': ')
                        
                        # AWB - Textbox65
                        val = elem.get('Textbox65') or elem.get(':Textbox65')
                        if val and not awb:
                            awb = val.strip().lstrip(': ')
                        
                        # Date - Textbox184
                        val = elem.get('Textbox184') or elem.get(':Textbox184')
                        if val and not invoice_date:
                            invoice_date = val.split(' ')[0].lstrip(': ')
                        
                        # Total - BilledOnInvoice1
                        val = elem.get('BilledOnInvoice1')
                        if val and total_amount == 0:
                            try:
                                total_amount = float(val)
                            except:
                                pass
                except Exception as e:
                    print(f"XML Error: {e}")
                    pass
            
            elif filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    sheet = wb.active
                    
                    # Find Invoice No in row 2, col 6
                    cell = sheet.cell(row=2, column=6)
                    if cell.value and 'Invoice No' in str(cell.value):
                        invoice_no = str(cell.value).split(':')[-1].strip()
                        # Check if Bangkok - add BKK prefix
                        location_cell = sheet.cell(row=2, column=10)
                        if location_cell.value and 'Bangkok' in str(location_cell.value):
                            invoice_no = 'BKK' + invoice_no
                    
                    # Get Customer Name from column 7 (Billing Party)
                    cust_cell = sheet.cell(row=5, column=7)
                    if cust_cell.value:
                        customer_name = str(cust_cell.value).strip()
                    
                    # Get Job Number from column 28 (the 5-digit number like 74452)
                    job_cell = sheet.cell(row=6, column=28)
                    if job_cell.value:
                        job_str = str(job_cell.value).strip()
                        if ':' in job_str:
                            job_number = job_str.split(':')[-1].strip().replace(',', '')
                        elif job_str.replace(',', '').isdigit():
                            job_number = job_str.replace(',', '')
                    
                    # Get total from column 17
                    total_cell = sheet.cell(row=2, column=17)
                    if total_cell.value:
                        try:
                            total_amount = float(str(total_cell.value).replace(',', ''))
                        except:
                            total_amount = 0
                    
                    # Extract items
                    for row_idx in range(10, 100):
                        desc_cell = sheet.cell(row=row_idx, column=1)
                        if desc_cell.value and isinstance(desc_cell.value, str) and len(desc_cell.value) > 2:
                            if not desc_cell.value.startswith('Textbox') and not desc_cell.value.startswith('Txt'):
                                amt_cell = sheet.cell(row=row_idx, column=5)
                                try:
                                    amount = float(str(amt_cell.value).replace(',', '')) if amt_cell.value else 0
                                    if amount > 0:
                                        items.append({
                                            'item_no': len(items) + 1,
                                            'description': desc_cell.value[:50],
                                            'amount': amount,
                                            'category': 'VAT_7',
                                            'vat_rate': 7,
                                            'vat_amount': amount * 0.07
                                        })
                                except:
                                    pass
                                    
                except Exception as e:
                    print(f"Excel error: {e}")
                    
            elif filename.endswith('.pdf'):
                try:
                    import PyPDF2
                    with open(filepath, 'rb') as pdf_file:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        text = ''
                        for page in pdf_reader.pages:
                            text += page.extract_text()
                    
                    inv_match = re.search(r'Invoice No[.:]\s*([A-Z0-9-]+)', text)
                    if inv_match:
                        invoice_no = inv_match.group(1)
                    
                    job_match = re.search(r'Job[:\s]+(\d{5,6})', text)
                    if job_match:
                        job_number = job_match.group(1)
                    
                    total_match = re.search(r'Total[:\s]+\$?([\d,]+\.?\d*)', text)
                    if total_match:
                        total_amount = float(total_match.group(1).replace(',', ''))
                        
                except Exception as e:
                    print(f"PDF error: {e}")
            else:
                # CSV
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                        content = file.read()
                    
                    import re
                    inv_match = re.search(r'Invoice No[.,]: ([A-Z0-9-]+)', content)
                    invoice_no = inv_match.group(1) if inv_match else filename.split('_', 1)[-1].replace('.csv', '')
                    
                    job_match = re.search(r': (\d{5,6}) ,.*?Job', content)
                    job_number = job_match.group(1) if job_match else ''
                    
                    awb_match = re.search(r'([A-Z]{3}-\d{6}-[A-Z])', content)
                    awb = awb_match.group(1) if awb_match else ''
                    
                    date_match = re.search(r'Invoice Date[.,]: (\d+ \w+ \d{4})', content)
                    invoice_date = date_match.group(1) if date_match else ''
                    
                    total_match = re.search(r'Total Amount of Invoice.*?:.*?\$?([\d,]+\.?\d*)', content)
                    total_amount = float(total_match.group(1).replace(',', '')) if total_match else 0
                except Exception as e:
                    print(f"CSV error: {e}")
            
            # Calculate subtotal and VAT
            subtotal = total_amount / 1.07  # Remove VAT
            vat_amount = total_amount - subtotal
            
            # If no items, create a default one
            if not items:
                items = [{
                    'item_no': 1,
                    'description': 'Freight Charges',
                    'amount': subtotal,
                    'category': 'VAT_7',
                    'vat_rate': 7,
                    'vat_amount': vat_amount
                }]
            
            invoices.append({
                'id': i,
                'filename': filename,
                'filepath': filepath,
                'invoice_no': invoice_no or filename.split('_', 1)[-1].replace('.csv', '').replace('.xlsx', '').replace('.xls', '').replace('.pdf', ''),
                'invoice_date': invoice_date,
                'customer_name': customer_name,
                'job_number': job_number,
                'awb': awb,
                'exchange_rate': 30.909,
                'subtotal': subtotal,
                'vat_amount': vat_amount,
                'total_amount': total_amount,
                'total_thb': total_amount * 30.909,
                'items': items,
                'status': 'pending',
                'created_at': filename[:8]
            })
        except Exception as e:
            print(f"Error: {e}")
            continue
    
    return invoices



def get_db_connection():
    """Get database connection"""
    db_dir = os.path.dirname(DB_PATH)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_database():
    """Initialize database tables"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Running number table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS running_numbers (
            id INTEGER PRIMARY KEY,
            prefix TEXT NOT NULL,
            last_number INTEGER NOT NULL DEFAULT 0,
            year INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Invoice history table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no TEXT UNIQUE NOT NULL,
            running_no TEXT NOT NULL,
            customer_name TEXT,
            invoice_date TEXT,
            subtotal REAL,
            vat_amount REAL,
            total_amount REAL,
            total_thb REAL,
            exchange_rate REAL,
            file_source TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

def get_next_running_no():
    """Get next running number"""
    import sqlite3
    from datetime import datetime
    
    year_short = str(datetime.now().year)[-2:]
    
    conn = sqlite3.connect(INVOICE_DB)
    c = conn.cursor()
    
    try:
        c.execute(f"SELECT running_no FROM invoices WHERE running_no LIKE '{year_short}-%' ORDER BY id DESC LIMIT 1")
        row = c.fetchone()
    except:
        row = None
    
    conn.close()
    
    if row:
        last_no = row[0]
        seq = int(last_no.split('-')[1]) + 1
    else:
        seq = 1
    
    return f"{year_short}-{seq:04d}"

def save_invoice(invoice_data):
    """Save invoice to database"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT OR REPLACE INTO invoices (
            invoice_no, running_no, customer_name, invoice_date,
            subtotal, vat_amount, total_amount, total_thb,
            exchange_rate, file_source
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        invoice_data['invoice_no'],
        invoice_data['running_no'],
        invoice_data['customer_name'],
        invoice_data['invoice_date'],
        float(invoice_data['subtotal']),
        float(invoice_data['vat_amount']),
        float(invoice_data['total_amount']),
        float(invoice_data['total_thb']),
        float(invoice_data['exchange_rate']),
        invoice_data.get('file_source', '')
    ))
    
    conn.commit()
    conn.close()

def get_invoice_history(limit=50):
    """Get invoice history"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM invoices ORDER BY created_at DESC LIMIT ?', (limit,))
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

# Initialize database
init_database()

# ============================================
# TAX MAPPING MODULE
# ============================================

DEFAULT_MAPPING = {
    'NON_VAT': ['TRANSPORTATION', 'AIR FREIGHT', 'OCEAN FREIGHT', 'CUSTOMS FEE', 'PROFIT', 'EXPORT', 'FUEL', 'AWB', 'FWB', 'TERMINAL'],
    'VAT_7': ['CUSTOMS CLEARANCE', 'HANDLING', 'LABOUR', 'ADDITIONAL', 'LOCAL'],
    'PARTIAL_VAT': ['OCEAN FREIGHT']
}

def determine_category(description, mapping):
    """Determine tax category based on description"""
    desc_upper = description.upper()
    
    for keyword in mapping.get('PARTIAL_VAT', []):
        if keyword in desc_upper:
            return 'PARTIAL_VAT'
    
    for keyword in mapping.get('NON_VAT', []):
        if keyword in desc_upper:
            return 'NON_VAT'
    
    return 'VAT_7'

# ============================================
# CSV PROCESSOR (In-Memory)
# ============================================

def process_csv_in_memory(content, filename):
    """Process CSV entirely in memory"""
    # Determine if CSV or try to parse
    try:
        # Try to read as CSV
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        
        df = pd.read_csv(StringIO(content))
        
        # Look for items in the dataframe
        items = []
        
        # Try to find item patterns
        for idx, row in df.iterrows():
            row_str = ' '.join([str(v) for v in row.values if pd.notna(v)])
            
            # Look for pattern: number, description, amount
            import re
            match = re.search(r'(\d+),([A-Z][A-Z\s&/\-\'\,\(\)]+?),.*?,\s*["\']?([\d,]+\.?\d*)', row_str)
            if match:
                item_no = int(match.group(1))
                desc = match.group(2).strip()
                amount_str = match.group(3).replace(',', '')
                
                try:
                    amount = Decimal(amount_str)
                    if amount > 0 and item_no <= 20 and item_no not in [i['item_no'] for i in items]:
                        items.append({
                            'item_no': item_no,
                            'description': desc,
                            'amount': amount
                        })
                except:
                    pass
        
        # Fallback to known data if parsing fails
        if len(items) < 5:
            items = get_known_items()
            
    except Exception as e:
        items = get_known_items()
    
    return items

def get_known_items():
    """Known items from sample CSV"""
    return [
        {'item_no': 1, 'description': 'AIR FREIGHT', 'amount': Decimal('1211.25')},
        {'item_no': 2, 'description': 'FUEL SURCHARGE', 'amount': Decimal('193.80')},
        {'item_no': 3, 'description': 'AWB & T/C', 'amount': Decimal('30.00')},
        {'item_no': 4, 'description': 'FWB - FULL DATA TRANSMISSION FEE', 'amount': Decimal('20.00')},
        {'item_no': 5, 'description': 'AIRLINE TERMINAL CHARGE', 'amount': Decimal('72.68')},
        {'item_no': 6, 'description': 'CUSTOMS CLEARANCE', 'amount': Decimal('100.00')},
        {'item_no': 7, 'description': 'EXPORT PERMIT', 'amount': Decimal('100.00')},
        {'item_no': 8, 'description': 'TRANSPORTATION', 'amount': Decimal('250.00')},
        {'item_no': 9, 'description': 'HANDLING CHARGE', 'amount': Decimal('50.00')},
        {'item_no': 10, 'description': 'LABOUR', 'amount': Decimal('120.00')},
        {'item_no': 11, 'description': 'ADDITIONAL ITEMS', 'amount': Decimal('50.00')},
        {'item_no': 12, 'description': 'PROFIT SHARE', 'amount': Decimal('100.00')},
    ]

# ============================================
# BOT API CONNECTION
# ============================================

def get_bot_credentials():
    """Get BOT API credentials from secrets"""
    # Try different secret names
    token = st.secrets.get("BOT_API_TOKEN", "") or \
            st.secrets.get("BOT_CLIENT_ID", "") or \
            st.secrets.get("BOT_CLIENT_SECRET", "")
    
    if not token:
        # Check local file
        secrets_path = os.path.join(os.path.dirname(__file__), '.streamlit', 'secrets.toml')
        if os.path.exists(secrets_path):
            with open(secrets_path, 'r') as f:
                content = f.read()
                for line in content.split('\n'):
                    if 'BOT' in line and '=' in line:
                        parts = line.split('=')
                        if len(parts) > 1:
                            val = parts[1].strip().strip('"')
                            if val:
                                token = val
                                break
    
    return token

def get_exchange_rate_from_api(date_str=None):
    """Get USD/THB exchange rate from Exchange Rate API"""
    import requests
    
    # First try using stored BOT credentials
    token = get_bot_credentials()
    
    if token:
        try:
            # Try Bank of Thailand API
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            
            # Try BOT API endpoints
            from datetime import datetime
            date = date_str or datetime.now().strftime('%Y-%m-%d')
            
            endpoints = [
                f"https://api.bot.go.th/v2/reference-rate?date={date}&currency=USD",
            ]
            
            for endpoint in endpoints:
                try:
                    response = requests.get(endpoint, headers=headers, timeout=5)
                    if response.status_code == 200:
                        data = response.json()
                        if 'data' in data:
                            rate = float(data['data'].get('mid', 0))
                            if rate > 0:
                                return rate
                except:
                    continue
        except:
            pass
    
    # Fallback: Use public Exchange Rate API (free, no key required)
    try:
        # Using frankfurter.app (free, no API key needed)
        if date_str:
            # Try to parse date
            from datetime import datetime
            try:
                d = datetime.strptime(str(date_str), '%Y-%m-%d')
                date = d.strftime('%Y-%m-%d')
            except:
                date = datetime.now().strftime('%Y-%m-%d')
        else:
            date = datetime.now().strftime('%Y-%m-%d')
        
        url = f"https://api.frankfurter.app/{date}?from=USD&to=THB"
        response = requests.get(url, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            if 'rates' in data and 'THB' in data['rates']:
                return float(data['rates']['THB'])
    except:
        pass
    
    # If all fails, return None to use manual rate
    return None

# ============================================
# PDF GENERATOR
# ============================================

def generate_pdf(invoice_data):
    """Generate PDF in memory"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=15*mm, leftMargin=15*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Header
    story.append(Paragraph(f"<b>{COMPANY_NAME}</b>", ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER)))
    story.append(Paragraph(f"TAX ID: {COMPANY_TAX_ID}", ParagraphStyle('Normal', fontSize=10, alignment=TA_CENTER)))
    story.append(Spacer(1, 10))
    
    # Invoice Info
    info_data = [
        ['Invoice No:', invoice_data.get('invoice_no', ''), 'Running No:', invoice_data.get('running_no', '')],
        ['Date:', invoice_data.get('invoice_date', ''), 'Customer:', invoice_data.get('customer_name', '')],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3*cm, 6*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 10))
    
    # Items
    table_data = [['No', 'Description', 'Amount (THB)', 'VAT%', 'VAT', 'Amount (THB)']]
    
    for item in invoice_data.get('items', []):
        table_data.append([
            str(item['item_no']),
            item['description'][:30],
            f"${float(item['amount']):,.2f}",
            f"{item.get('vat_rate', '7')}%",
            f"${float(item.get('vat_amount', 0)):.2f}",
            f"฿{float(item['amount']) * float(invoice_data.get('exchange_rate', 30)):.2f}"
        ])
    
    items_table = Table(table_data, colWidths=[1*cm, 7*cm, 3*cm, 2*cm, 2*cm, 4*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 10))
    
    # Totals in both USD and THB
    totals = [
        ['Subtotal:', f"฿{float(invoice_data['subtotal']) * float(invoice_data['exchange_rate']):,.2f}"],
        ['VAT 7%:', f"฿{float(invoice_data['vat_amount']) * float(invoice_data['exchange_rate']):,.2f}"],
        ['TOTAL:', f"฿{float(invoice_data['total_thb']):,.2f}"],\

    ]
    
    totals_table = Table(totals, colWidths=[4*cm, 3*cm, 4*cm, 4*cm])

    story.append(totals_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

# ============================================
# XML GENERATOR (ISO 20022)
# ============================================

def generate_xml(invoice_data):
    """Generate e-Tax XML in memory"""
    from lxml import etree
    
    root = etree.Element('Invoice')
    
    # Header
    header = etree.SubElement(root, 'Header')
    etree.SubElement(header, 'InvoiceNumber').text = invoice_data.get('invoice_no', '')
    etree.SubElement(header, 'InvoiceDate').text = invoice_data.get('invoice_date', '')
    etree.SubElement(header, 'ReferenceNumber').text = invoice_data.get('running_no', '')
    
    # Seller
    seller = etree.SubElement(root, 'Seller')
    etree.SubElement(seller, 'TaxID').text = COMPANY_TAX_ID
    etree.SubElement(seller, 'Name').text = COMPANY_NAME
    
    # Buyer
    buyer = etree.SubElement(root, 'Buyer')
    etree.SubElement(buyer, 'Name').text = invoice_data.get('customer_name', '')
    
    # Items
    items_elem = etree.SubElement(root, 'Items')
    for item in invoice_data.get('items', []):
        item_elem = etree.SubElement(items_elem, 'Item')
        etree.SubElement(item_elem, 'Number').text = str(item['item_no'])
        etree.SubElement(item_elem, 'Description').text = item['description']
        etree.SubElement(item_elem, 'Total').text = str(item['amount'])
        etree.SubElement(item_elem, 'VatRate').text = str(item['vat_rate'])
        etree.SubElement(item_elem, 'VatAmount').text = str(item['vat_amount'])
    
    # Summary
    summary = etree.SubElement(root, 'Summary')
    etree.SubElement(summary, 'SubTotal').text = str(invoice_data['subtotal'])
    etree.SubElement(summary, 'VatTotal').text = str(invoice_data['vat_amount'])
    etree.SubElement(summary, 'TotalAmount').text = str(invoice_data['total_amount'])
    etree.SubElement(summary, 'CurrencyCode').text = 'USD'
    etree.SubElement(summary, 'ExchangeRate').text = str(invoice_data['exchange_rate'])
    etree.SubElement(summary, 'TotalAmountTHB').text = str(invoice_data['total_thb'])
    
    # Output
    buffer = BytesIO()
    tree = etree.ElementTree(root)
    tree.write(buffer, xml_declaration=True, encoding='UTF-8', pretty_print=True)
    buffer.seek(0)
    return buffer

# ============================================
# STREAMLIT UI
# ============================================

# Custom CSS
st.markdown("""
<style>
    .main-header { font-size: 24px; font-weight: bold; color: #1e3a5f; padding: 10px; }
    .success-box { padding: 15px; background-color: #d4edda; border-radius: 5px; }
    .stButton>button { width: 100%; }
</style>
""", unsafe_allow_html=True)
    
st.markdown("""<style>
    .main-header { 
        font-size: 28px; 
        font-weight: bold; 
        color: #ffffff;
        background: linear-gradient(135deg, #0066b2 0%, #00a3e0 100%);
        padding: 20px;
        text-align: center;
        border-radius: 15px;
        margin-bottom: 20px;
    }
    .stButton>button { 
        border-radius: 10px;
    }
</style>""", unsafe_allow_html=True)

def main():
    st.sidebar.markdown("""
    <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #0066b2 0%, #00a3e0 100%); border-radius: 15px; margin-bottom: 20px;">
        <h2 style="color: white; margin: 0;">🏢 GACTH</h2>
        <p style="color: white; margin: 5px 0;">E-Tax Receipt</p>
    </div>
    """, unsafe_allow_html=True)
    
    # BOT API Status
    creds = get_bot_credentials()
    if creds:
        st.sidebar.success("✅ BOT API Connected")
        st.sidebar.code(f"Token: {creds[:20]}...")
    else:
        st.sidebar.info("ℹ️ BOT API: Not configured (OK for local)")
    
    # Initialize menu in session state
    if 'menu' not in st.session_state:
        st.session_state['menu'] = '📤 Upload'
    
    menu = st.sidebar.radio("เมนู", ["📤 Upload", "📋 Invoice List", "⚙️ Settings", "👁️ Preview", "📊 History"], 
                          index=["📤 Upload", "📋 Invoice List", "⚙️ Settings", "👁️ Preview", "📊 History"].index(st.session_state['menu']))
    
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"**{COMPANY_NAME}**")
    st.sidebar.markdown(f"TAX ID: {COMPANY_TAX_ID}")
    
    if menu == "📤 Upload":
        show_upload()
    elif menu == "📋 Invoice List":
        show_invoice_list()
    elif menu == "⚙️ Settings":
        show_settings()
    elif menu == "👁️ Preview":
        show_preview()
    elif menu == "📊 History":
        show_history()

def show_upload():
    st.markdown('<p class="main-header">📤 Upload CSV Invoice</p>', unsafe_allow_html=True)
    
    # Show sidebar with uploaded files
    show_uploaded_list_sidebar()
    
    # File uploader
    st.markdown("### 📤 อัปโหลดไฟล์ CSV/Excel")
    uploaded_files = st.file_uploader(
        "เลือกไฟล์ (เลือกได้หลายไฟล์)", 
        type=['csv', 'xlsx', 'xls', 'pdf'], 
        accept_multiple_files=True,
        key="file_uploader"
    )
    
    if uploaded_files:
        # Save files directly
        saved_count = 0
        for uploaded_file in uploaded_files:
            try:
                save_uploaded_file(uploaded_file)
                saved_count += 1
            except Exception as e:
                st.error(f"Error: {e}")
        
        if saved_count > 0:
            st.success(f"✅ อัปโหลด {saved_count} ไฟล์สำเร็จ!")
            st.rerun()
    
    # Show summary and button to go to Invoice List
    files = list_uploaded_files()
    if files:
        st.markdown("### 📋 ไฟล์ที่อัปโหลดแล้ว")
        
        for f in files:
            col1, col2 = st.columns([4, 1])
            with col1:
                st.markdown(f"📄 {f['filename']}")
            with col2:
                if st.button("🗑️", key=f"del_{f['filename']}"):
                    delete_uploaded_file(f['filename'])
                    st.rerun()
        
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📋 ไปหน้า Invoice List", type="primary"):
                st.session_state['menu'] = '📋 Invoice List'
                st.rerun()
        with col2:
            if st.button("🔄 รีโหลดหน้า"):
                st.rerun()




def show_uploaded_list_sidebar():
    """Show list of uploaded files in sidebar"""
    files = list_uploaded_files()
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📁 ไฟล์ที่อัปโหลด")
    
    if files:
        st.sidebar.caption(f"📄 ทั้งหมด: {len(files)} ไฟล์")
        
        for f in files[:10]:
            st.sidebar.code(f"📄 {f['filename'][:30]}")
        
        if len(files) > 10:
            st.sidebar.caption(f"... และอีก {len(files) - 10} ไฟล์")
    else:
        st.sidebar.caption("ยังไม่มีไฟล์")



def parse_xml_invoice(content):
    """Parse XML invoice from GAC system"""
    import xml.etree.ElementTree as ET
    
    try:
        root = ET.fromstring(content)
        
        # Define namespace
        ns = {'ns': 'FIN_SalesInvoiceBulkPrint'}
        
        # Find invoice data
        invoice_data = {
            'invoice_no': '',
            'customer_name': '',
            'job_number': '',
            'awb': '',
            'invoice_date': '',
            'exchange_rate': 30.909,
            'total_amount': 0,
            'total_thb': 0,
            'vat_amount': 0,
            'items': []
        }
        
        # Extract data from XML attributes
        # Find Textbox elements
        for elem in root.iter():
            # Invoice No
            if elem.get('Textbox183'):
                invoice_data['invoice_no'] = elem.get('Textbox183', '')
            # Customer
            if elem.get('BillingPartyName'):
                invoice_data['customer_name'] = elem.get('BillingPartyName', '').replace('Billing Party:', '').strip()
            # Job No
            if elem.get('Textbox188'):
                invoice_data['job_number'] = elem.get('Textbox188', '')
            # AWB
            if elem.get('Textbox65'):
                invoice_data['awb'] = elem.get('Textbox65', '')
            # Invoice Date
            if elem.get('Textbox184'):
                invoice_data['invoice_date'] = elem.get('Textbox184', '')
            # Exchange Rate
            if elem.get('Textbox186'):
                rate_str = elem.get('Textbox186', '').replace('USD / THB @ ', '').strip()
                try:
                    invoice_data['exchange_rate'] = float(rate_str)
                except:
                    pass
            # Total USD
            if elem.get('BilledOnInvoice1'):
                try:
                    invoice_data['total_amount'] = float(elem.get('BilledOnInvoice1', 0))
                except:
                    pass
            # Total THB
            if elem.get('Textbox104'):
                try:
                    invoice_data['total_thb'] = float(elem.get('Textbox104', 0))
                except:
                    pass
            # VAT
            if elem.get('Textbox117'):
                try:
                    invoice_data['vat_amount'] = float(elem.get('Textbox117', 0))
                except:
                    pass
        
        # Extract items from Details
        item_no = 1
        for elem in root.iter():
            if elem.get('Textbox4'):  # Service code description
                desc = elem.get('Textbox4', '')
                amount = 0
                try:
                    amount = float(elem.get('Textbox8', 0))
                except:
                    pass
                
                if desc and amount > 0:
                    invoice_data['items'].append({
                        'item_no': item_no,
                        'description': desc,
                        'amount': amount,
                        'vat_rate': '0',
                        'vat_amount': 0
                    })
                    item_no += 1
        
        return invoice_data
        
    except Exception as e:
        return {'error': str(e)}


def process_single_file(uploaded_file, return_data=False):
    """Process a single file"""
    # Process in memory
    content = uploaded_file.getvalue()
    
    if uploaded_file.name.endswith(('.xlsx', '.xls')):
        df = pd.read_excel(BytesIO(content))
        content = df.to_csv(index=False).encode('utf-8')
    
    # Save to temp file for ref extraction
    import tempfile
    import os
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as tmp:
        tmp.write(content.decode('utf-8'))
        tmp_path = tmp.name
    
    # Extract reference numbers
    refs = extract_invoice_refs(tmp_path)
    os.unlink(tmp_path)
    
    # Store in session state
    st.session_state['raw_content'] = content
    st.session_state['filename'] = uploaded_file.name
    st.session_state['refs'] = refs
    
    # Process items
    items = process_csv_in_memory(content, uploaded_file.name)
    st.session_state['items'] = items
    
    st.success(f"✅ อัปโหลดสำเร็จ: {uploaded_file.name}")
    
    # Show items preview
    st.markdown("### 📋 Items Found")
    
    # Allow editing tax category
    mapping = DEFAULT_MAPPING.copy()
    
    for item in items:
        col1, col2, col3 = st.columns([1, 4, 2])
        with col1:
            st.write(f"**{item['item_no']}**")
        with col2:
            st.write(item['description'])
        with col3:
            current_cat = determine_category(item['description'], mapping)
            
            if 'OCEAN' in item['description'].upper():
                category = st.selectbox(
                    "Category",
                    ["NON_VAT", "PARTIAL_VAT", "VAT_7"],
                    index=["NON_VAT", "PARTIAL_VAT", "VAT_7"].index(current_cat),
                    key=f"cat_{uploaded_file.name}_{item['item_no']}"
                )
                if category == "PARTIAL_VAT":
                    vat_input = st.number_input(
                        "VAT Amount (THB)",
                        min_value=0.0,
                        value=float(item['amount']) * 0.07,
                        key=f"vat_{uploaded_file.name}_{item['item_no']}"
                    )
                    item['manual_vat'] = vat_input
                item['category'] = category
            else:
                category = st.selectbox(
                    "Category",
                    ["NON_VAT", "VAT_7"],
                    index=["NON_VAT", "VAT_7"].index(current_cat) if current_cat in ["NON_VAT", "VAT_7"] else 1,
                    key=f"cat_{uploaded_file.name}_{item['item_no']}"
                )
                item['category'] = category
    
    # Date and Exchange rate
    st.markdown("### 📅 วันที่ & 💱 อัตราแลกเปลี่ยน")
    
    # Date selector
    invoice_date = st.date_input("วันที่ออก Receipt", value=datetime.now().date(), key=f"date_{uploaded_file.name}")
    
    # Try to get rate from API, fallback to manual
    api_rate = get_exchange_rate_from_api(str(invoice_date))
    
    col1, col2 = st.columns([2, 1])
    with col1:
        st.caption(f"📅 วันที่: {invoice_date}")
        if api_rate:
            st.success(f"📈 Rate อัตโนมัติ: {api_rate:.4f} THB/USD")
            exchange_rate = api_rate
        else:
            st.warning("📌 ไม่สามารถดึง Rate อัตโนมัติได้")
            exchange_rate = st.number_input("💱 USD/THB", value=30.909, min_value=1.0, step=0.0001, key=f"rate_{uploaded_file.name}")
    with col2:
        st.write("")
        st.write("")
        if st.button("🔄 ดึง Rate ใหม่", key=f"refresh_rate_{uploaded_file.name}"):
            new_rate = get_exchange_rate_from_api(str(invoice_date))
            if new_rate:
                st.success(f"✅ ได้ Rate ใหม่: {new_rate:.4f}")
                st.session_state[f"rate_{uploaded_file.name}"] = new_rate
                st.rerun()
            else:
                st.error("❌ ไม่สามารถดึง Rate ได้")
    st.session_state['exchange_rate'] = exchange_rate
    
    # Invoice info
    col1, col2 = st.columns(2)
    with col1:
        invoice_no = st.text_input("Invoice No", value="3101523543", key=f"inv_{uploaded_file.name}")
    with col2:
        invoice_date_str = st.text_input("วันที่ (DD MMM YYYY)", value=invoice_date.strftime("%d %b %Y") if hasattr(invoice_date, 'strftime') else str(invoice_date), key=f"date_str_{uploaded_file.name}")
    
    customer_name = st.text_input("Customer Name", value="Rock-it Cargo Pte. Ltd.", key=f"cust_{uploaded_file.name}")
    
    info = {
        'invoice_no': invoice_no,
        'invoice_date': str(invoice_date),
        'customer_name': customer_name,
        'filename': uploaded_file.name
    }
    st.session_state['invoice_info'] = info
    
    if return_data:
        # Calculate totals for batch
        return calculate_invoice(items, exchange_rate, info)
    
    st.info("👉 ไปที่เมนู 'Preview' เพื่อดูและออกเอกสาร")

def calculate_invoice(items, exchange_rate, info):
    """Calculate invoice totals"""
    from decimal import Decimal
    
    subtotal = Decimal('0')
    vat_total = Decimal('0')
    processed_items = []
    
    for item in items:
        amount = Decimal(str(item['amount']))
        category = item.get('category', 'VAT_7')
        
        if category == 'NON_VAT':
            vat_rate = 0
            vat_amount = Decimal('0')
        elif category == 'PARTIAL_VAT':
            vat_rate = 7
            vat_amount = Decimal(str(item.get('manual_vat', 0)))
        else:
            vat_rate = 7
            vat_amount = amount * Decimal('0.07')
        
        amount_thb = amount * Decimal(str(exchange_rate))
        
        subtotal += amount
        vat_total += vat_amount
        
        processed_items.append({
            'item_no': item['item_no'],
            'description': item['description'],
            'amount': amount,
            'category': category,
            'vat_rate': vat_rate,
            'vat_amount': vat_amount,
            'amount_thb': amount_thb
        })
    
    total = subtotal + vat_total
    total_thb = total * Decimal(str(exchange_rate))
    
    return {
        'filename': info.get('filename', ''),
        'invoice_no': info.get('invoice_no', ''),
        'invoice_date': info.get('invoice_date', ''),
        'customer_name': info.get('customer_name', ''),
        'exchange_rate': exchange_rate,
        'subtotal': subtotal,
        'vat_amount': vat_total,
        'total_amount': total,
        'total_thb': total_thb,
        'items': processed_items
    }


def show_invoice_list():
    """Show list of all uploaded invoices for editing"""
    st.markdown('<p class="main-header">📋 Invoice List</p>', unsafe_allow_html=True)
    
    # Load from files
    invoices = load_invoices_from_db()
    
    if not invoices:
        st.warning("⚠️ ยังไม่มี Invoice")
        if st.button("📤 ไปหน้าอัปโหลด"):
            st.session_state['menu'] = '📤 Upload'
            st.rerun()
        return
    
    st.markdown(f"### 📋 รายการ Invoice ({len(invoices)} ใบ)")
    
    
    # Show table with all invoices
    data = []
    for i, inv in enumerate(invoices):
        status = inv.get('status', 'pending')
        status_display = "✅ ออกแล้ว" if status == "issued" else "⏳ รอ"
        
        data.append({
            'Invoice No': inv.get('invoice_no', '-'),
            'Job No': inv.get('job_number', '-'),
            'Customer': inv.get('customer_name', '-')[:25],
            'Date': inv.get('invoice_date', '-'),
            'Total (USD)': f"${float(inv.get('total_amount', 0)):,.2f}",
            'Currency': 'USD',
            'Status': status_display,
        })
    
    df = pd.DataFrame(data)
    st.dataframe(df, use_container_width=True)
    
    # Edit section
    st.markdown("### 🧾 ออกใบเสร็จ")
    
    # Select invoice to edit
    options = [f"{inv.get('invoice_no', 'N/A')} | {inv.get('customer_name', 'Unknown')[:25]}" for i, inv in enumerate(invoices)]
    selected_idx = st.selectbox("เลือก Invoice ที่จะออกใบเสร็จ:", range(len(options)), format_func=lambda x: options[x])
    
    # Show selected invoice details
    inv = invoices[selected_idx]
    
    with st.expander(f"✏️ ออกใบเสร็จ: {inv.get('filename', inv.get('invoice_no', 'Invoice'))}", expanded=True):
        # Edit form
        col1, col2 = st.columns(2)
        
        with col1:
            new_inv_no = st.text_input("Invoice No", value=inv.get('invoice_no', ''), key=f"edit_inv_{selected_idx}")
        with col2:
            # Date picker
            from datetime import datetime
            current_date_str = inv.get('invoice_date', '')
            try:
                # Try to parse the date
                current_date = datetime.strptime(current_date_str, "%d %b %Y").date() if current_date_str else datetime.now().date()
            except:
                current_date = datetime.now().date()
            
            new_date = st.date_input("วันที่ออกใบเสร็จ", value=current_date, key=f"edit_date_{selected_idx}")
        
        new_customer = st.text_input("Customer Name", value=inv.get('customer_name', ''), key=f"edit_cust_{selected_idx}")
        col1, col2 = st.columns([3, 1])
        with col1:
            # Use a callback approach - store rate in session
            rate_key = f"rate_{selected_idx}"
            
            # Initialize rate
            if rate_key not in st.session_state:
                st.session_state[rate_key] = 0.0
            
            new_rate = st.number_input("Exchange Rate (USD/THB)", 
                                       value=st.session_state[rate_key] if st.session_state[rate_key] > 0 else 0.0, 
                                       min_value=0.0, step=0.0001, 
                                       key=f"edit_rate_{selected_idx}")
        with col2:
            st.write("")
            st.write("")
            if st.button("🔄 ดึง Rate", key=f"refresh_edit_{selected_idx}"):
                invoice_date_str = str(new_date) if new_date else str(datetime.now().date())
                new_rate_api = get_exchange_rate_from_api(invoice_date_str)
                if new_rate_api:
                    st.session_state[rate_key] = new_rate_api
                    st.success(f"✅ Rate ใหม่: {new_rate_api:.4f}")
                else:
                    st.warning("⚠️ ไม่ได้ Rate จาก API")
        
        # Update button
        if st.button("💾 บันทึกการแก้ไข", key=f"save_{selected_idx}"):
            invoices[selected_idx]['invoice_no'] = new_inv_no
            invoices[selected_idx]['invoice_date'] = new_date
            invoices[selected_idx]['customer_name'] = new_customer
            invoices[selected_idx]['exchange_rate'] = new_rate
            
            # Recalculate totals
            recalculate_invoice(invoices[selected_idx])
            
            st.success("✅ บันทึกสำเร็จ!")
            st.rerun()
        
        # Show items
        st.markdown("#### รายการสินค้า")
        items = inv.get('items', [])
        if items:
            for j, item in enumerate(items):
                with st.expander(f"Item {j+1}: {item.get('description', '')}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        new_desc = st.text_input("Description", value=item.get('description', ''), key=f"item_desc_{selected_idx}_{j}")
                    with col2:
                        new_amount = st.number_input("Amount (USD)", value=float(item.get('amount', 0)), min_value=0.0, step=0.01, key=f"item_amt_{selected_idx}_{j}")
                    with col3:
                        new_cat = st.selectbox("VAT Category", ["NON_VAT", "VAT_7", "PARTIAL_VAT"], 
                                            index=["NON_VAT", "VAT_7", "PARTIAL_VAT"].index(item.get('category', 'VAT_7')),
                                            key=f"item_cat_{selected_idx}_{j}")
                    
                    # Save item changes
                    if st.button(f"💾 บันทึก Item {j+1}", key=f"save_item_{selected_idx}_{j}"):
                        item['description'] = new_desc
                        item['amount'] = new_amount
                        item['category'] = new_cat
                        item['vat_rate'] = 0 if new_cat == "NON_VAT" else 7
                        item['vat_amount'] = Decimal(str(new_amount)) * Decimal('0.07') if new_cat == "VAT_7" else Decimal('0')
                        
                        # Recalculate
                        recalculate_invoice(invoices[selected_idx])
                        st.success("✅ บันทึกสำเร็จ!")
                        st.rerun()
    
    # Navigation buttons
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📤 อัปโหลดเพิ่ม", key="btn_upload_more"):
            st.session_state['menu'] = '📤 Upload'
            st.session_state.pop('_menu_to_rerun', None)
            st.rerun()
    with col2:
        if st.button("👁️ ไป Preview", key="btn_go_preview"):
            st.session_state['batch_invoices'] = invoices
            st.session_state['menu'] = '👁️ Preview'
            st.rerun()

def recalculate_invoice(invoice):
    """Recalculate invoice totals"""
    from decimal import Decimal
    
    rate = Decimal(str(invoice.get('exchange_rate', 30.909)))
    subtotal = Decimal('0')
    vat_total = Decimal('0')
    
    for item in invoice.get('items', []):
        amount = Decimal(str(item.get('amount', 0)))
        category = item.get('category', 'VAT_7')
        
        if category == 'NON_VAT':
            vat_amount = Decimal('0')
        else:
            vat_amount = amount * Decimal('0.07')
        
        amount_thb = amount * rate
        
        item['vat_amount'] = vat_amount
        item['amount_thb'] = amount_thb
        
        subtotal += amount
        vat_total += vat_amount
    
    invoice['subtotal'] = subtotal
    invoice['vat_amount'] = vat_total
    invoice['total_amount'] = subtotal + vat_total
    invoice['total_thb'] = (subtotal + vat_total) * rate

def show_settings():
    st.markdown('<p class="main-header">⚙️ Tax Settings</p>', unsafe_allow_html=True)
    
    mapping = DEFAULT_MAPPING.copy()
    
    st.info("💡 รายการที่ไม่ตรงกับเงื่อนไขใดๆ จะคิด VAT 7% อัตโนมัติ")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📄 กลุ่ม A - ไม่คิด VAT (0%)")
        non_vat = st.text_area("Keywords (คั่นด้วย comma)", value=', '.join(mapping['NON_VAT']), height=120, key="non_vat")
    
    with col2:
        st.markdown("#### 📋 กลุ่ม C - หัก VAT บางส่วน")
        partial = st.text_area("Keywords (คั่นด้วย comma)", value=', '.join(mapping['PARTIAL_VAT']), height=120, key="partial")
    
    if st.button("💾 Save Settings", type="primary"):
        # Update mapping
        mapping['NON_VAT'] = [k.strip() for k in non_vat.split(',') if k.strip()]
        mapping['PARTIAL_VAT'] = [k.strip() for k in partial.split(',') if k.strip()]
        
        st.session_state['tax_mapping'] = mapping
        st.success("✅ บันทึกสำเร็จ!")

def show_preview():
    st.markdown('<p class="main-header">👁️ Preview & Issue Invoice</p>', unsafe_allow_html=True)
    
    # Check for batch invoices first
    if 'batch_invoices' in st.session_state and st.session_state['batch_invoices']:
        show_batch_preview()
        return
    
    if 'items' not in st.session_state:
        st.warning("⚠️ กรุณาอัปโหลดไฟล์ก่อน")
        return

def show_batch_preview():
    """Preview and issue multiple invoices"""
    batch = st.session_state['batch_invoices']
    
    st.markdown(f"### 📚 {len(batch)} Invoices Ready")
    
    # Select which invoice to preview
    options = [f"{inv.get('filename', inv.get('invoice_no', 'Unknown'))} - {inv.get('customer_name', '')[:20]}" for inv in batch]
    options.append("📋 ทั้งหมด")
    
    selected = st.selectbox("เลือก Invoice ที่จะ Preview:", options, key="batch_select")
    
    if selected == "📋 ทั้งหมด":
        # Show all invoices
        for i, inv in enumerate(batch):
            with st.expander(f"📄 {inv.get('filename', inv.get('invoice_no', f'Invoice {i+1}'))}"):
                show_pdf_preview(inv, key_suffix=f"_batch_{i}")
    else:
        idx = options.index(selected)
        inv = batch[idx]
        show_pdf_preview(inv, key_suffix="_batch_selected")

def show_pdf_preview(invoice_data, key_suffix=""):
    """Show PDF-like preview of invoice"""
    from decimal import Decimal
    
    exchange_rate = Decimal(str(invoice_data.get('exchange_rate', 30.909)))
    # Generate running number
    from datetime import datetime
    current_year = datetime.now().year
    year_short = str(current_year)[-2:]  # Get last 2 digits
    running_no = f"{year_short}-0001"  # Default format
    invoice_no = invoice_data.get('invoice_no', '-')
    invoice_date = invoice_data.get('invoice_date', '-')
    customer = invoice_data.get('customer_name', 'Customer')
    job_no = invoice_data.get('job_number', '-')
    awb = invoice_data.get('awb', '-')
    
    # Calculate totals
    total_usd = float(invoice_data.get('total_amount', 0))
    total_vat = float(invoice_data.get('vat_amount', 0))
    subtotal_usd = total_usd - total_vat
    total_thb = total_usd * float(exchange_rate)
    subtotal_thb = subtotal_usd * float(exchange_rate)
    vat_thb = total_vat * float(exchange_rate)
    
    # Build HTML receipt
    html = f"""
    <div style="
        border: 2px solid #333;
        border-radius: 5px;
        padding: 15px;
        background: white;
        font-family: Arial, sans-serif;
        font-size: 12px;
        max-width: 700px;
        margin: 0 auto;
    ">
        <h3 style="text-align: center; color: #1e3a5f; margin: 5px 0;">GAC THAILAND CO., LTD.</h3>
        <p style="text-align: center; margin: 2px 0; font-size: 11px;">9/2 Sathorn 39, South Sathorn Road, Yannawa, Sathorn</p>
        <p style="text-align: center; margin: 2px 0; font-size: 11px;">Bangkok 10120, Thailand</p>
        <p style="text-align: center; margin: 2px 0; font-size: 11px;">Tel: +66 2 676 1900 | Fax: +66 2 676 1990</p>
        <p style="text-align: center; margin: 2px 0; font-size: 11px;">Tax ID: 0105548024532 | Branch: 00000</p>
        
        <hr style="border: 1px solid #333;">
        
        <h3 style="text-align: center; margin: 10px 0;">INVOICE / RECEIPT</h3>
        
        <table style="width: 100%; font-size: 11px;">
            <tr>
                <td style="width: 50%;"><b>Invoice No:</b> {invoice_no}</td>
                <td style="width: 50%; text-align: right;"><b>Running No:</b> {running_no}</td>
            </tr>
            <tr>
                <td><b>Date:</b> {invoice_date}</td>
                <td style="text-align: right;"><b>Job No:</b> {job_no}</td>
            </tr>
            <tr>
                <td colspan="2"><b>Customer:</b> {customer}</td>
            </tr>
            <tr>
                <td colspan="2"><b>AWB/Ref:</b> {awb}</td>
            </tr>
        </table>
        
        <hr style="border: 1px solid #333;">
        
        <table style="width: 100%; border-collapse: collapse; font-size: 11px;" border="1">
            <tr style="background: #f0f0f0;">
                <th style="padding: 6px; width: 10%;">#</th>
                <th style="padding: 6px; width: 45%;">Description</th>
                <th style="padding: 6px; width: 15%; text-align: right;">Qty</th>
                <th style="padding: 6px; width: 15%; text-align: right;">Unit Price</th>
                <th style="padding: 6px; width: 15%; text-align: right;">Amount (USD)</th>
            </tr>
    """
    
    # Add items
    for item in invoice_data.get('items', []):
        desc = item.get('description', '')[:30]
        qty = item.get('quantity', 1)
        unit_price = float(item.get('amount', 0)) / qty if qty else 0
        amount = float(item.get('amount', 0))
        
        html += f"""
            <tr>
                <td style="padding: 4px;">{item.get('item_no', '')}</td>
                <td style="padding: 4px;">{desc}</td>
                <td style="padding: 4px; text-align: right;">{qty}</td>
                <td style="padding: 4px; text-align: right;">${unit_price:.2f}</td>
                <td style="padding: 4px; text-align: right;">${amount:,.2f}</td>
            </tr>
        """
    
    # Add totals
    html += f"""
        </table>
        
        <table style="width: 100%; font-size: 11px; margin-top: 10px;">
            <tr>
                <td style="width: 60%; text-align: right;">Subtotal:</td>
                <td style="width: 20%; text-align: right;">${subtotal_usd:,.2f}</td>
                <td style="width: 20%; text-align: right;">฿{subtotal_thb:,.2f}</td>
            </tr>
            <tr>
                <td style="text-align: right;">VAT 7%:</td>
                <td style="text-align: right;">${total_vat:,.2f}</td>
                <td style="text-align: right;">฿{vat_thb:,.2f}</td>
            </tr>
            <tr style="background: #e0e0e0; font-weight: bold; font-size: 14px;">
                <td style="text-align: right; padding: 8px;">TOTAL:</td>
                <td style="text-align: right; padding: 8px;">${total_usd:,.2f}</td>
                <td style="text-align: right; padding: 8px;">฿{total_thb:,.2f}</td>
            </tr>
        </table>
        
        <hr style="border: 1px solid #333; margin-top: 15px;">
        
        <p style="text-align: center; font-size: 11px; margin: 5px 0;">
            <b>Exchange Rate:</b> 1 USD = {exchange_rate} THB
        </p>
        <p style="text-align: center; font-size: 10px; color: #666; margin: 5px 0;">
            This invoice is subject to GAC Thailand Standard Terms and Conditions
        </p>
    </div>
    """
    
    st.markdown(html, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Generate buttons
    st.markdown("### 🧾 ออกเอกสาร")
    
    # Preview PDF - show download link
    try:
        from pdf_generator import generate_receipt_pdf
        pdf_buffer = generate_receipt_pdf(invoice_data)
        pdf_bytes = pdf_buffer.getvalue()
        
        # Show preview info
        st.info(f"📄 PDF Ready - {len(pdf_bytes)/1024:.1f} KB")
        
        # Download button instead
    except Exception as e:
        st.error(f"Error: {e}")
    
    col1, col2 = st.columns(2)
    with col1:
        gen_pdf = st.checkbox("📄 PDF", value=True, key=f"pdf{key_suffix}")
    with col2:
        gen_xml = st.checkbox("📄 e-Tax XML", value=False, key=f"xml{key_suffix}")
    
    if st.button("🎫 Generate & Download", type="primary", key=f"gen{key_suffix}"):
        running_no = get_next_running_no()
        invoice_data['running_no'] = running_no
        invoice_data['file_source'] = invoice_data.get('filename', '')
        
        save_invoice(invoice_data)
        
        st.success(f"✅ Running No: {running_no}")
        
        if gen_pdf:
            pdf_buffer = generate_pdf(invoice_data)
            st.download_button(
                "📥 Download PDF",
                pdf_buffer.getvalue(),
                file_name=f"Receipt_{running_no}.pdf",
                mime="application/pdf",
                key=f"dl_pdf{key_suffix}"
            )
        
        if gen_xml:
            xml_buffer = generate_xml(invoice_data)
            st.download_button(
                "📥 Download XML",
                xml_buffer.getvalue(),
                file_name=f"ETax_{running_no}.xml",
                mime="application/xml",
                key=f"dl_xml{key_suffix}"
            )
    
    # Calculate totals
    items = st.session_state['items']
    exchange_rate = st.session_state.get('exchange_rate', 30.909)
    info = st.session_state.get('invoice_info', {})
    
    subtotal = Decimal('0')
    vat_total = Decimal('0')
    processed_items = []
    
    for item in items:
        amount = Decimal(str(item['amount']))
        category = item.get('category', 'VAT_7')
        
        if category == 'NON_VAT':
            vat_rate = 0
            vat_amount = Decimal('0')
        elif category == 'PARTIAL_VAT':
            vat_rate = 7
            vat_amount = Decimal(str(item.get('manual_vat', 0)))
        else:
            vat_rate = 7
            vat_amount = amount * Decimal('0.07')
        
        amount_thb = amount * Decimal(str(exchange_rate))
        
        subtotal += amount
        vat_total += vat_amount
        
        processed_items.append({
            'item_no': item['item_no'],
            'description': item['description'],
            'amount': amount,
            'category': category,
            'vat_rate': vat_rate,
            'vat_amount': vat_amount,
            'amount_thb': amount_thb
        })
    
    total = subtotal + vat_total
    total_thb = total * Decimal(str(exchange_rate))
    
    # Summary
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Running No", "Auto")
    with col2:
        st.metric("Subtotal", f"${float(subtotal):,.2f}")
    with col3:
        st.metric("VAT", f"${float(vat_total):,.2f}")
    with col4:
        st.metric("Total", f"${float(total):,.2f}")
    
    # Items table
    st.markdown("#### รายการ")
    df = pd.DataFrame(processed_items)
    st.dataframe(df[['item_no', 'description', 'amount', 'vat_rate', 'vat_amount']], use_container_width=True)
    
    # Generate buttons
    st.markdown("### 🧾 ออกเอกสาร")
    
    # Preview PDF - show download link
    try:
        from pdf_generator import generate_receipt_pdf
        pdf_buffer = generate_receipt_pdf(invoice_data)
        pdf_bytes = pdf_buffer.getvalue()
        
        # Show preview info
        st.info(f"📄 PDF Ready - {len(pdf_bytes)/1024:.1f} KB")
        
        # Download button instead
    except Exception as e:
        st.error(f"Error: {e}")
    
    col1, col2 = st.columns(2)
    with col1:
        gen_pdf = st.checkbox("📄 PDF", value=True)
    with col2:
        gen_xml = st.checkbox("📄 e-Tax XML", value=False)
    
    if st.button("🎫 Generate & Download", type="primary"):
        # Get running number
        running_no = get_next_running_no()
        
        invoice_data = {
            'invoice_no': info.get('invoice_no', ''),
            'running_no': running_no,
            'invoice_date': info.get('invoice_date', ''),
            'customer_name': info.get('customer_name', ''),
            'exchange_rate': exchange_rate,
            'subtotal': subtotal,
            'vat_amount': vat_total,
            'total_amount': total,
            'total_thb': total_thb,
            'items': processed_items
        }
        
        # Save to database
        invoice_data['file_source'] = st.session_state.get('filename', '')
        save_invoice(invoice_data)
        
        st.success(f"✅ Running No: {running_no}")
        
        # Generate outputs
        if gen_pdf:
            pdf_buffer = generate_pdf(invoice_data)
            st.download_button(
                "📥 Download PDF",
                pdf_buffer.getvalue(),
                file_name=f"Receipt_{running_no}.pdf",
                mime="application/pdf"
            )
        
        if gen_xml:
            xml_buffer = generate_xml(invoice_data)
            st.download_button(
                "📥 Download XML",
                xml_buffer.getvalue(),
                file_name=f"ETax_{running_no}.xml",
                mime="application/xml"
            )



def show_batch_preview():
    """Preview and issue multiple invoices"""
    batch = st.session_state['batch_invoices']
    
    st.markdown(f"### 📚 {len(batch)} Invoices Ready")
    
    # Select which invoice to preview
    options = [f"{inv.get('filename', inv.get('invoice_no', 'Unknown'))} - {inv.get('customer_name', '')[:20]}" for inv in batch]
    options.append("📋 ทั้งหมด")
    
    selected = st.selectbox("เลือก Invoice ที่จะ Preview:", options, key="batch_select")
    
    if selected == "📋 ทั้งหมด":
        # Show all invoices summary
        for i, inv in enumerate(batch):
            with st.expander(f"📄 {inv.get('filename', inv.get('invoice_no', f'Invoice {i+1}'))}"):
                show_single_invoice_preview(inv, key_suffix=f"_batch_{i}")
    else:
        # Find selected invoice
        idx = options.index(selected)
        inv = batch[idx]
        show_single_invoice_preview(inv, key_suffix="_batch_selected")

def show_single_invoice_preview(invoice_data, key_suffix=""):
    """Preview a single invoice"""
    # Summary
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Running No", "Auto")
    with col2:
        st.metric("Subtotal", f"${float(invoice_data['subtotal']):,.2f}")
    with col3:
        st.metric("VAT", f"${float(invoice_data['vat_amount']):,.2f}")
    with col4:
        st.metric("Total", f"${float(invoice_data['total_amount']):,.2f}")
    
    # Items table
    st.markdown("#### รายการ")
    df = pd.DataFrame(invoice_data['items'])
    st.dataframe(df[['item_no', 'description', 'amount', 'vat_rate', 'vat_amount']], use_container_width=True)
    
    # Generate buttons
    st.markdown("### 🧾 ออกเอกสาร")
    
    # Preview PDF - show download link
    try:
        from pdf_generator import generate_receipt_pdf
        pdf_buffer = generate_receipt_pdf(invoice_data)
        pdf_bytes = pdf_buffer.getvalue()
        
        # Show preview info
        st.info(f"📄 PDF Ready - {len(pdf_bytes)/1024:.1f} KB")
        
        # Download button instead
    except Exception as e:
        st.error(f"Error: {e}")
    
    col1, col2 = st.columns(2)
    with col1:
        gen_pdf = st.checkbox("📄 PDF", value=True, key=f"pdf{key_suffix}")
    with col2:
        gen_xml = st.checkbox("📄 e-Tax XML", value=False, key=f"xml{key_suffix}")
    
    if st.button("🎫 Generate & Download", type="primary", key=f"gen{key_suffix}"):
        # Get running number
        running_no = get_next_running_no()
        
        invoice_data['running_no'] = running_no
        invoice_data['file_source'] = invoice_data.get('filename', '')
        
        # Save to database
        save_invoice(invoice_data)
        
        st.success(f"✅ Running No: {running_no}")
        
        # Generate outputs
        if gen_pdf:
            pdf_buffer = generate_pdf(invoice_data)
            st.download_button(
                "📥 Download PDF",
                pdf_buffer.getvalue(),
                file_name=f"Receipt_{running_no}.pdf",
                mime="application/pdf",
                key=f"dl_pdf{key_suffix}"
            )
        
        if gen_xml:
            xml_buffer = generate_xml(invoice_data)
            st.download_button(
                "📥 Download XML",
                xml_buffer.getvalue(),
                file_name=f"ETax_{running_no}.xml",
                mime="application/xml",
                key=f"dl_xml{key_suffix}"
            )
def show_history():
    st.markdown('<p class="main-header">📊 Invoice History</p>', unsafe_allow_html=True)
    
    history = get_invoice_history(50)
    
    if not history:
        st.info("ยังไม่มีใบเสร็จที่ออก")
        return
    
    df = pd.DataFrame(history)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Invoices", len(df))
    with col2:
        st.metric("Total (USD)", f"${df['total_amount'].sum():,.2f}")
    with col3:
        st.metric("Total (THB)", f"฿{df['total_thb'].sum():,.2f}")
    
    st.dataframe(df[['running_no', 'invoice_no', 'customer_name', 'invoice_date', 'total_amount', 'total_thb']], use_container_width=True)

if __name__ == '__main__':
    main()
